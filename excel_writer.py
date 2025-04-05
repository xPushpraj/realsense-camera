from openpyxl import Workbook, load_workbook
import os

def save_to_excel(name, measurements, filename="measurements.xlsx"):
    headers = ['Name', 'Max Reach', 'Chest Diameter', 'Bicep Diameter', 'Thigh to Feet', 'Full Height']
    
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Data'
        ws.append(headers)
    else:
        wb = load_workbook(filename)
        ws = wb['Data']

    ws.append([name] + measurements)
    wb.save(filename)
